import os
import sys

import asyncio

import tkinter

from tkinter import filedialog

import fnmatch

from autobahn.asyncio.wamp import ApplicationRunner
from ak_autobahn import AkComponent

from waapi import WAAPI_URI





class MyComponent(AkComponent):


    WwiseProjectPath = ""
    pathToWwiseProject = ""
    OriginalsPathRoot = ""







    def onJoin(self, details):

############# Function definitions #########################################
        def exit():
            self.leave()

        def beginUndoGroup():
            yield from self.call(WAAPI_URI.ak_wwise_core_undo_begingroup)

        def cancelUndoGroup():
            yield from self.call(WAAPI_URI.ak_wwise_core_undo_cancelgroup)

        def endUndoGroup():
            undoArgs = {"displayName": "My Waapi Script"}
            yield from self.call(WAAPI_URI.ak_wwise_core_undo_endgroup, {}, **undoArgs)

        def saveWwiseProject():
            yield from self.call(WAAPI_URI.ak_wwise_core_project_save)

        def askUserForAudioDirectory():
            root = tkinter.Tk()
            root.withdraw()
            root.update()
            AudioDir = filedialog.askdirectory(title="Select Originals Folder")
            root.update()
            root.destroy()
            return AudioDir


        def askUserForExcelFile():
            root = tkinter.Tk()
            root.withdraw()
            root.update()
            ExcelFile = filedialog.askopenfilename(initialdir="/", title="Select Excel Doc",
                                                   filetypes=(("ExcelDoc", "*.xlsx",), ("all files", "*.*")))
            root.update()
            root.destroy()
            return ExcelFile




        def checkIfDirectoryIsPopulated(Path):

            if [f for f in os.listdir(Path) if not f.startswith('.')] == []:
                #print(Path + " is not populated")
                FolderPopulated = False
            else:
                #print(Path + " is  populated")
                FolderPopulated = True

            print(os.listdir(Path))
            return FolderPopulated

        def createFolder(Path):

            if not os.path.exists(Path):
                os.makedirs(Path)
                #print(cellvalue + "Folder Was Created")
                FolderExisted = True
            else:
                #print(cellvalue + "Folder Already Existed")
                FolderExisted = False

            return FolderExisted




        def createWwiseObject(args):
            try:
                res = yield from self.call(WAAPI_URI.ak_wwise_core_object_create, {}, **args)
            except Exception as ex:
                print("call error: {}".format(ex))






        def SetObjectReference(args):
            try:
                res = yield from self.call(WAAPI_URI.ak_wwise_core_object_setreference, {}, **args)
            except Exception as ex:
                print(cellvalue + " Object Reference Already Set".format(ex))


        def SetObjectProperty(args):
            try:
                res = yield from self.call(WAAPI_URI.ak_wwise_core_object_setproperty, {}, **args)
            except Exception as ex:
                print(cellvalue + " Object Property Already Set".format(ex))

        def AssignChildtoSwitch(args):
            try:
                yield from self.call(WAAPI_URI.ak_wwise_core_switchcontainer_addassignment, {}, **args)
            except Exception as ex:
                pass
               #print(cellvalue + " Already Assigned To Switch".format(ex))

        def importAudioFiles(args):
            try:
                yield from self.call(WAAPI_URI.ak_wwise_core_audio_import, {}, **args)
                #print("audiofilesimported")
            except Exception as ex:
                print("call error: {}".format(ex))



        def setupCreateArgs(parentID ,otype = "BlendContainer",  oname = "" , conflict = "merge"):
            setSwitchArgs = {

                "parent": parentID,
                "type": otype,
                "name": oname,
                "onNameConflict": "merge"
            }

            return setSwitchArgs

        def setupAssignSwitchGroupArgs (objectID, SwitchPath, rtype = "SwitchGroupOrStateGroup"):
            createObjArgs = {

                "object": objectID,
                "reference": rtype,
                "value": SwitchPath
            }

            return createObjArgs


        def setupAssignChildToSwitch( childcontainer, switch):

            ChildSwitchAssignArgs = {

                "child": childcontainer,
                "stateOrSwitch": switch,

            }

            return ChildSwitchAssignArgs


        def setupInclusionArgs (objectID, shouldbeincluded, rtype = "Inclusion"):
            createObjArgs = {

                "object": objectID,
                "property": rtype,
                "value": shouldbeincluded
            }

            return createObjArgs


        def setupColorArgs (objectID, colorvalueInt):
            ColorArgs = {

                "object": objectID,
                "property": "Color",
                "value": colorvalueInt
            }

            return ColorArgs

        def EventCreateArgs(parentID,EventName, EventTargetPath, action = 1):

            createObjArgs = {

                "parent": parentID,
                "type": "Event",
                "name": EventName,
                "notes": "Action 0 can be overwritten by script -  if you wish to override which container is played, or add additional actions, Uninclude Action 0 and start fromm Action 1",
                "onNameConflict": "merge",
                "children": [
                {   "name": "0",
                    "type": "Action",
                    "@ActionType": action,
                    "@Target": EventTargetPath},
                ]
            }

            return createObjArgs


        def SetupIntObjContainerCreateArgs(parentID, ContainerName):
            createObjArgs = {

                "parent": parentID,
                "type": "BlendContainer",
                "name": ContainerName,
                "onNameConflict": "merge",
                "children": [
                    {"name": ContainerName,
                     "type": "RandomSequenceContainer"},
                    {"name": ContainerName + "_Sweetener",
                     "type": "RandomSequenceContainer"},
                ]
            }
            return createObjArgs


        def SetupSwitchContainerCreateArgs(parentID, SwitchContainerName):

            createObjArgs = {

                "parent": parentID,
                "type": "SwitchContainer",
                "name": SwitchContainerName,
                "onNameConflict": "merge",
                ##"SwitchGroup": "\\Switches\\Physics\\PhysicsImpact_Velocity",
                ##"switchAssignation": "<PhysicsImpact_Velocity>",
                "children": [
                    {"name": SwitchContainerName + "_Soft",
                     "type": "RandomSequenceContainer"},
                    {"name": SwitchContainerName + "_Hard",
                     "type": "RandomSequenceContainer"},
                ]
            }
            return createObjArgs


        def SetupPhysicsImpactSwitchGroupCreateArgs(parentID, SwitchGroupName):

            createObjArgs = {

                "parent": parentID,
                "type": "SwitchGroup",
                "name": SwitchGroupName,
                "onNameConflict": "merge",
                "children": [
                    {"name": "Soft",
                     "type": "Switch"},
                    {"name": "Hard",
                     "type": "Switch"},
                ]
            }
            return createObjArgs


        def setupSourceFileList(path):
            # print("Setting up list of audio files")
            filelist = []
            pattern = '*.wav'

            for root, dirs, files in os.walk(path):
                # for file in os.listdir(path):
                for filename in fnmatch.filter(files, pattern):
                    absFilePath = os.path.abspath(os.path.join(root, filename))
                    filelist.append(absFilePath)
            return filelist

        def setupImportWavFileToContainer (AudioFile, ObjectName, ImportDestinationContainer):

            createImportArgs = {


                "importOperation": "replaceExisting",
                "autoAddToSourceControl": True,
                 "imports": [
                    {
                    "audioFile": AudioFile,
                    "importLocation" : ImportDestinationContainer,
                    "objectType": "Sound SFX",
                    "objectPath": ObjectName,
                    },
                ]
            }
            return createImportArgs


        def getWwiseObjects(args):
            # print("Get a list of the audio files currently in the project, under the selected object")

            try:
                res = yield from self.call(WAAPI_URI.ak_wwise_core_object_get, **args)
            except Exception as ex:
                print("call error: {}".format(ex))
                return 0
            else:
                return res.kwresults["return"]


        def getProject():
            arguments = {
                "from": {"ofType": ["Project"]},
                "options": {
                    "return": ["type","id", "name", "filePath"]
                }
            }
            try:
                res = yield from self.call(WAAPI_URI.ak_wwise_core_object_get, **arguments)
            except Exception as ex:
                print("call error: {}".format(ex))
                cancelUndoGroup()
            else:
                MyComponent.WwiseProjectPath = res.kwresults["return"][0]['filePath']
                MyComponent.WwiseProjectPath = MyComponent.WwiseProjectPath.replace("Y:","~").replace('\\', '/')
                MyComponent.pathToWwiseProject = os.path.dirname(os.path.abspath(MyComponent.WwiseProjectPath))
                MyComponent.OriginalsPathRoot = os.path.join(MyComponent.pathToWwiseProject, "Originals\\SFX").replace('\\', '/')


        def CheckIfAllDescendantsHaveAudioCDescendantsAndSetInclusion (path):

            arguments = {
                "from": {"path": [path]},
                "transform": [
                    {"select": ["descendants"]},
                    # {"where": ["type:isIn", ["Sound"]]}
                ],
                "options": {
                    "return": ["id", "type", "name", "path", ]
                }
            }

            # Get all the audio files from a Wwise path
            listofcontainers = yield from getWwiseObjects(arguments)

            for x in listofcontainers:


                if not x["type"] == "Sound" and not x["type"] == "AudioFileSource":
                    #
                    # print(x["name"])

                    #### for any descendants that are NOT sounds or audio file sources...see if they have sound or audio source descendants

                    arguments = {

                        "from": {"path": [x["path"]]},
                        "transform": [
                            {"select": ["descendants"]},
                            {"where": ["type:isIn", ["Sound"]]}
                        ],
                        "options": {
                            "return": ["name", ]
                        }
                    }

                    ContainerPath = x["path"]
                    # print (ContainerPath)
                    res = yield from getWwiseObjects(arguments)

                    #### set to uninclude if no audio descendants

                    if not res:
                        # print("exclude" + x["name"])

                        ShouldBeIncluded = False

                    else:
                        # print("Include" + x["name"])

                        ShouldBeIncluded = True

                    args = setupInclusionArgs(ContainerPath, ShouldBeIncluded)
                    yield from SetObjectProperty(args)

        def CheckIfSelectedContainerHasAudioDescendantsAndSetInclusion(path):

            arguments = {
                "from": {"path": [path]},
                "transform": [
                    {"select": ["descendants"]},
                    {"where": ["type:isIn", ["Sound"]]}
                ],

                 "options": {
                    "return": ["name"]
                }
            }


            res = yield from self.call(WAAPI_URI.ak_wwise_core_object_get, **arguments)

            #### set to uninclude if no audio descendants

            if not res.kwresults["return"]:



                ShouldBeIncluded = False


            else:

                ShouldBeIncluded = True

            args = setupInclusionArgs(path, ShouldBeIncluded)
            yield from SetObjectProperty(args)

            #print ("ShouldBeInclude = " + str(ShouldBeIncluded))
            return ShouldBeIncluded


        def GetListOfAudioFilesInSelectedContainer(path, ContainerName = "containername"):


            arguments = {
                "from": {"path": [path]},
                "transform": [
                    {"select": ["descendants"]},
                    {"where": ["type:isIn", ["Sound"]]}
                ],

                 "options": {
                    "return": ["name"]
                }
            }

            res = yield from self.call(WAAPI_URI.ak_wwise_core_object_get, **arguments)
            filelist = []

            for f in res.kwresults["return"]:
                fileName = str(f["name"])
                filelist.append(fileName)

            print(ContainerName + " File list:")
            print(filelist)
            return filelist


        ############## End of Function definitions ##############################################



############### Main Script logic flow ####################################################
        ## First try to connect to wwise to get Wwise info
        try:
            res = yield from self.call(WAAPI_URI.ak_wwise_core_getinfo)  # RPC call without arguments
        except Exception as ex:
            print("call error: {)".format(ex))
        else:
            # Call was successful, displaying information from the payload.
            print("Hello {} {}".format(res.kwresults['displayName'], res.kwresults['version']['displayName']))




    ###########  Do Some Cool stuff here ##############

        beginUndoGroup()

        yield from getProject()



        #############   Ask User for Excel Document #####

        SourceExcelFile = "Whiteboard-AudioMaterials-GrabsandImpactsChecklist.xlsx"
        #SourceExcelFile = os.path.expanduser(askUserForExcelFile())
        #print(SourceExcelFile)

        #############   Ask User for Audio Directory  #####

        AudioDir = MyComponent.OriginalsPathRoot
        #AudioDir = os.path.expanduser(askUserForAudioDirectory())
        print(AudioDir)

        #############   Load Excel Doc ############

        try:
            from openpyxl import load_workbook  # Excel import
        except:
            print("Need openpyxl to open Excel file, installing...")
            os.system('py -m pip install openpyxl')
            from openpyxl import load_workbook  # Excel import

        ExcelFile = SourceExcelFile
        wb = load_workbook(ExcelFile)
        page = wb['Object Recording Track List']
        print(page)


       #############  Setup main Hierarchy and targets ########



################ Start Physics Set up  #######################

 ########### Create Physics Top work unit, actor mixer, Impact actor mixer and Impact velocity Switch Group  #####


    ####### Top Physics Structure  #####


        PhysicsAudioPath = "\\Actor-Mixer Hierarchy\\Whiteboard\\Physics\\"

        args = setupCreateArgs(PhysicsAudioPath, "WorkUnit", "Physics")
        yield from createWwiseObject(args)

        args = setupCreateArgs(PhysicsAudioPath + "Physics\\", "ActorMixer", "Physics")
        yield from createWwiseObject(args)

        PhysicsAudioPath = "\\Actor-Mixer Hierarchy\\Whiteboard\\Physics\\Physics\\Physics\\"


        args = setupCreateArgs(PhysicsAudioPath , "ActorMixer", "Impact")
        yield from createWwiseObject(args)


        PhysicsImpactPath = PhysicsAudioPath + "Impact\\"


        PhysicsImpactAudioSourceDir = AudioDir + "/Physics/Impacts/"


        PhysicsEventPath = "\\Events\\Whiteboard\\Physics\\"


        args = setupCreateArgs(PhysicsEventPath, "WorkUnit", "Physics")
        yield from createWwiseObject(args)


        PhysicsEventPath = "\\Events\\Whiteboard\\Physics\\Physics\\"


        args = setupCreateArgs(PhysicsEventPath, "Folder", "Impact")
        yield from createWwiseObject(args)

        PhysicsImpactEventPath = PhysicsEventPath + "Impact\\"


    ####### Create Impact Velocity Switch Group ####


        args = setupCreateArgs("\\Switches\\Whiteboard\\Whiteboard\\", "WorkUnit", "Physics")
        yield from createWwiseObject(args)


        args = SetupPhysicsImpactSwitchGroupCreateArgs("\\Switches\\Whiteboard\\Whiteboard\\Physics\\","PhysicsImpact_Velocity")
        yield from createWwiseObject(args)

        PhysicsImpactVeloctySwitchPath = "\\Switches\\Whiteboard\\Whiteboard\\Physics\\PhysicsImpact_Velocity"

        PhysicsImpactEventNamePrepend = "Play_SFX_WB_Physics_Impact_"

        ########Do everything twice, once for objects once for surfaces#######




        MaterialTypeList = ["Object", "Surface"]


        for x in MaterialTypeList:



            PhysicsImpactPath = PhysicsAudioPath + "Impact\\"

            PhysicsImpactAudioSourceDir = AudioDir + "/Physics/Impacts/"

            MaterialType = x

            args = setupCreateArgs(PhysicsImpactPath, "ActorMixer", MaterialType)
            yield from createWwiseObject(args)

            PhysicsImpactPath = PhysicsImpactPath + MaterialType + "\\"

            PhysicsImpactAudioSourceDir = PhysicsImpactAudioSourceDir + MaterialType + "/"




            args = setupCreateArgs(PhysicsEventPath + "Impact\\", "Folder", MaterialType)
            yield from createWwiseObject(args)




    ######## Reset Excel Counters ######

            material = []
            rowStart = 3
            columnStart = 2
            rowCounter = 0

           ## beginUndoGroup()

            page = wb[x + ' Recording Track List']
            print(page)


            #### Start creating OBJECT MATERIAL impacts structures######

            while (True):


                cellvalue = page.cell(row=rowStart + rowCounter, column=columnStart).value
                if (cellvalue == None):
                    rowCounter = 0
                    break
                rowCounter += 1
                material.append(cellvalue)

                TargetContainer = cellvalue
                ShouldCreateEvents = True
                ShouldBeIncluded = True

                createFolder(PhysicsImpactAudioSourceDir + cellvalue + "/" + cellvalue + "_Soft")
                createFolder(PhysicsImpactAudioSourceDir + cellvalue + "/" + cellvalue + "_Hard")

                #     TargetContainer = "_Default"
                #     ShouldBeIncluded = False
                # #
                # if not (checkIfDirectoryIsPopulated(ObjectImpactsAudioDir + cellvalue + "/" + cellvalue + "_Soft")) and not (checkIfDirectoryIsPopulated(ObjectImpactsAudioDir + cellvalue + "/" + cellvalue + "_Hard")):
                #     # if the folder did exist, see if its empty, if it is don't do anything as it will have already been set to default intially but may have been overidden in wwise for events to target somewhere else ###
                #     # if the folder is not empty, set the target to its own containers'
                #
                #     ShouldCreateEvents = False
                #     ShouldBeIncluded = False


                    #print(cellvalue + " Object Impacts  Already existed but was empty so event wasnt created")

                args = SetupSwitchContainerCreateArgs(PhysicsImpactPath, cellvalue)
                yield from createWwiseObject(args)


                # args = setupInclusionArgs(PhysicsImpactPath + cellvalue, ShouldBeIncluded)
                # yield from SetObjectProperty(args)


                ################ set up event for Object Impact Switch container ####


                ImpactTypeList = ["Soft" , "Hard"]

                for ImpactType in ImpactTypeList:

                    AudioSourceDir = PhysicsImpactAudioSourceDir + cellvalue + "/" + cellvalue + "_" + ImpactType

                    ImportFileList = setupSourceFileList(AudioSourceDir)

                    ImportDestinationContainer = PhysicsImpactPath + cellvalue + "\\" + cellvalue + "_" + ImpactType

                    AudioFilesInSelectedContainer = yield from GetListOfAudioFilesInSelectedContainer(ImportDestinationContainer, "Impact " + MaterialType + " " + cellvalue + " " + ImpactType )

                    for file in ImportFileList:
                        f = file.rsplit('.')
                        fname = os.path.basename(f[0])
                        AudioFileName = fname

                        AudioFile = os.path.abspath(file)

                        #print(AudioFileName)

                        if AudioFileName not in AudioFilesInSelectedContainer:

                            print("New File To Import: - " + AudioFileName)

                            args = setupImportWavFileToContainer(AudioFile, AudioFileName, ImportDestinationContainer)
                            yield from importAudioFiles(args)

                            print("imported audio file for " + cellvalue + "_Impacts_" + MaterialType + "_" + ImpactType)
                        #else:
                            #print("File Already Imported")
                    print(" ")

                ################ set up  Children and Swithc Assignments for Object Impact Switch container ####

                args = setupAssignSwitchGroupArgs(PhysicsImpactPath + cellvalue, PhysicsImpactVeloctySwitchPath)
                yield from SetObjectReference(args)

                args = setupAssignSwitchGroupArgs(PhysicsImpactPath + cellvalue, PhysicsImpactVeloctySwitchPath + "\\Soft", "DefaultSwitchOrState")
                yield from SetObjectReference(args)

                args = setupAssignChildToSwitch( PhysicsImpactPath + cellvalue + "\\" + cellvalue + "_Soft", PhysicsImpactVeloctySwitchPath +"\\Soft")
                yield from AssignChildtoSwitch(args)

                args = setupAssignChildToSwitch(PhysicsImpactPath + cellvalue + "\\" + cellvalue + "_Hard", PhysicsImpactVeloctySwitchPath +"\\Hard")
                yield from AssignChildtoSwitch(args)

    ###################   Set up Events ############

                if not (yield from CheckIfSelectedContainerHasAudioDescendantsAndSetInclusion(PhysicsImpactPath + cellvalue)):

                    TargetContainer = "_Default"



                EventPath = PhysicsImpactEventPath + MaterialType
                EventName = PhysicsImpactEventNamePrepend + MaterialType + "_" + cellvalue

                TargetPath = PhysicsImpactPath + TargetContainer

                args = EventCreateArgs(EventPath, EventName, TargetPath)
                yield from createWwiseObject(args)

                # args = setupColorArgs(EventPath + "\\" + EventName, 1)
                # yield from SetObjectProperty(args)



                ##print(cellvalue + "Object Impacts Done")
            print("ALL " +MaterialType + " Impacts Setup Done!")

            #input("Press any key to exit")



        # endUndoGroup()
        # saveWwiseProject()




        exit()




############### Exit  #############################




    def onDisconnect(self):
        print("The client was disconnected.")

        asyncio.get_event_loop().stop()


if __name__ == '__main__':
    runner = ApplicationRunner(url=u"ws://127.0.0.1:8095/waapi", realm=u"realm1")

    try:
        runner.run(MyComponent)
    except Exception as e:
        print(type(e).__name__ + ": Is Wwise running and Wwise Authoring API enabled?")
